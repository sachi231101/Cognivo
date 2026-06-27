"""Text extraction helpers for PDF, DOCX, TXT, XLSX."""
import io
from typing import Optional

from pypdf import PdfReader
from docx import Document as DocxDoc
import openpyxl


def extract_text(filename: str, content: bytes) -> str:
    lower = filename.lower()
    try:
        if lower.endswith(".pdf"):
            return _pdf(content)
        if lower.endswith(".docx"):
            return _docx(content)
        if lower.endswith((".txt", ".md")):
            return content.decode("utf-8", errors="ignore")
        if lower.endswith((".xlsx", ".xlsm")):
            return _xlsx(content)
        # fallback: try decoding as plain text
        return content.decode("utf-8", errors="ignore")
    except Exception as e:
        return f"[Could not parse {filename}: {e}]"


def _pdf(content: bytes) -> str:
    reader = PdfReader(io.BytesIO(content))
    parts = []
    for page in reader.pages:
        try:
            parts.append(page.extract_text() or "")
        except Exception:
            continue
    return "\n".join(parts).strip()


def _docx(content: bytes) -> str:
    doc = DocxDoc(io.BytesIO(content))
    return "\n".join(p.text for p in doc.paragraphs if p.text).strip()


def _xlsx(content: bytes) -> str:
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    parts = []
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        parts.append(f"### Sheet: {sheet}")
        for row in ws.iter_rows(values_only=True):
            cells = [str(c) for c in row if c is not None]
            if cells:
                parts.append(" | ".join(cells))
    return "\n".join(parts).strip()


def chunk_text(text: str, chunk_size: int = 1200, overlap: int = 150) -> list[str]:
    """Split text into overlapping word-aware chunks."""
    if not text:
        return []
    words = text.split()
    chunks = []
    step = max(1, chunk_size - overlap)
    i = 0
    while i < len(words):
        chunk = " ".join(words[i : i + chunk_size])
        if chunk.strip():
            chunks.append(chunk)
        i += step
    return chunks


def score_chunk(chunk: str, query_keywords: list[str]) -> int:
    """Simple keyword scoring."""
    lower = chunk.lower()
    return sum(lower.count(k) for k in query_keywords if k)
